from __future__ import annotations

import re

import openpyxl

from core.result import AgentResult

_HOJAS_OBLIGATORIAS = ["descriptivos", "modelo"]
_PREFIJO_BIVARIADO = "bivariado_"
_FILAS_REQUERIDAS = ["b", "ll", "ul"]

# Termino de interaccion de Stata de `mean ..., over()`: `c.{outcome}@{codigo}[bn].{pred}`.
# Solo interesa el codigo de categoria (el `bn` es la marca de nivel base de Stata).
_TERMINO_STATA = re.compile(r"@(\d+)(?:bn)?\.")


def _limpiar_termino(cn: str) -> str:
    """Convierte el nombre crudo de la matriz de Stata en una etiqueta legible.
    `c.higiene_oral@1bn.area` -> `1`. Los terminos del modelo (nombres de
    variable, `_cons`) no matchean el patron y se devuelven intactos."""
    m = _TERMINO_STATA.search(cn)
    return m.group(1) if m else cn


def _leer_hoja(ws) -> tuple[list[str], dict[str, dict[str, object]]]:
    # `putexcel ... = matrix(r(table)), names` escribe una fila extra de
    # ecuacion (nombre del outcome, repetido en cada columna) para estimadores
    # con ecuacion (ologit/mlogit/logistic/regress) -- pero no para `mean`. Si
    # la fila 1 es un solo valor repetido en mas de una columna, es esa fila
    # de ecuacion -- los terminos reales estan en la fila 2.
    fila1 = [c.value for c in ws[1][1:] if c.value is not None]
    header_row = 1
    col_names = fila1
    if len(fila1) > 1 and len(set(fila1)) == 1:
        fila2 = [c.value for c in ws[2][1:] if c.value is not None]
        if len(fila2) > 1 and len(set(fila2)) > 1:
            col_names = fila2
            header_row = 2
    filas: dict[str, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=header_row + 1):
        label = row[0].value
        if label is None:
            continue
        filas[str(label)] = {cn: row[i + 1].value for i, cn in enumerate(col_names)}
    return col_names, filas


def _tabla_terminos(col_names, filas) -> list[dict]:
    terminos = []
    for cn in col_names:
        termino = _limpiar_termino(cn)
        try:
            item = {
                "termino": termino,
                "efecto": float(filas["b"][cn]),
                "ic_inf": float(filas["ll"][cn]),
                "ic_sup": float(filas["ul"][cn]),
                "p": float(filas["pvalue"][cn]) if "pvalue" in filas else None,
            }
        except (TypeError, ValueError):
            if termino == "_cons":
                continue
            raise
        terminos.append(item)
    return terminos


def parsear_resultados(xlsx_path: str) -> AgentResult:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    faltan = [h for h in _HOJAS_OBLIGATORIAS if h not in wb.sheetnames]
    if faltan:
        return AgentResult.failure([f"resultados.xlsx: falta(n) hoja(s): {faltan}"])
    data: dict[str, list[dict]] = {}
    warnings: list[str] = []
    for hoja in _HOJAS_OBLIGATORIAS:
        col_names, filas = _leer_hoja(wb[hoja])
        for req in _FILAS_REQUERIDAS:
            if req not in filas:
                return AgentResult.failure(
                    [f"resultados.xlsx[{hoja}]: falta la fila requerida '{req}'"])
        try:
            data[hoja] = _tabla_terminos(col_names, filas)
        except (TypeError, ValueError):
            return AgentResult.failure(
                [f"resultados.xlsx[{hoja}]: celda no numérica donde se esperaba un número"])
    bivariado: dict[str, list[dict]] = {}
    for hoja in wb.sheetnames:
        if not hoja.startswith(_PREFIJO_BIVARIADO):
            continue
        pred = hoja[len(_PREFIJO_BIVARIADO):]
        col_names, filas = _leer_hoja(wb[hoja])
        if not all(req in filas for req in _FILAS_REQUERIDAS):
            warnings.append(f"resultados.xlsx[{hoja}]: bivariado mal formado, se omite")
            continue
        try:
            bivariado[pred] = _tabla_terminos(col_names, filas)
        except (TypeError, ValueError):
            warnings.append(f"resultados.xlsx[{hoja}]: celda no numérica en bivariado, se omite")
    data["bivariado"] = bivariado
    return AgentResult.success(data, warnings=warnings)
