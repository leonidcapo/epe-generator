import json
import shutil
from pathlib import Path

import pytest
import openpyxl

from core.knowledge import load_perfil, guardar_perfil, Perfil
from core.sheets_client import FakeSheetReader
from core.llm_client import FakeLLMClient
from core.pubmed_client import FakePubMedClient
from tests.fixtures.sheet_rows_sinteticas import FILAS_SINTETICAS


@pytest.fixture(autouse=True)
def _sin_llm_real(monkeypatch):
    """Prevent any test in this file from using a real LLM environment or loading a real .env.

    This autouse fixture ensures:
    - DEEPSEEK_API_KEY and LLM_PROVIDER are never loaded from the environment
    - load_dotenv() is monkeypatched to do nothing, preventing accidental .env discovery

    This closes the risk class where load_dotenv() in orchestrator.main() walks up the
    directory tree and loads a real .env with real API keys that leak into os.environ,
    affecting subsequent tests in the same pytest session.
    """
    monkeypatch.delenv("DEEPSEEK_API_KEY", raising=False)
    monkeypatch.delenv("LLM_PROVIDER", raising=False)
    monkeypatch.setattr("dotenv.load_dotenv", lambda *a, **k: None)
    yield


def _copiar_plantilla(tmp_path):
    repo_root = Path(__file__).resolve().parent.parent
    (tmp_path / "knowledge").mkdir(exist_ok=True)
    shutil.copy(repo_root / "knowledge" / "plantilla_epe.yaml",
                tmp_path / "knowledge" / "plantilla_epe.yaml")


def _copiar_limitaciones(tmp_path):
    repo_root = Path(__file__).resolve().parent.parent
    (tmp_path / "knowledge").mkdir(exist_ok=True)
    shutil.copy(repo_root / "knowledge" / "limitaciones_epe.yaml",
                tmp_path / "knowledge" / "limitaciones_epe.yaml")


def _resultados_xlsx_valido(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "descriptivos"
    ws.append([None, "nivel_tratamiento_requerido"])
    ws.append(["b", 2.34])
    ws.append(["ll", 2.10])
    ws.append(["ul", 2.58])
    ws2 = wb.create_sheet("modelo")
    ws2.append([None, "riesgo_sistemico_asa"])
    ws2.append(["b", 1.87])
    ws2.append(["ll", 1.20])
    ws2.append(["ul", 2.91])
    path = tmp_path / "resultados.xlsx"
    wb.save(path)
    return str(path)


def test_run_perfilar_ok_escribe_perfil(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    reader = FakeSheetReader(FILAS_SINTETICAS)
    r = orchestrator.run_perfilar(reader, out_path="knowledge/perfil_epe.yaml")
    assert r.ok
    assert (tmp_path / "knowledge" / "perfil_epe.yaml").exists()


def test_run_perfilar_fallo_usa_cache_si_existe(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    cache_path = tmp_path / "knowledge" / "perfil_epe.yaml"
    guardar_perfil(Perfil(n_por_celda={("adultos", "riesgo_sistemico_asa"): 99},
                          distribuciones={}, generado_en="2026-07-01"), str(cache_path))
    reader = FakeSheetReader([], fail=True)
    r = orchestrator.run_perfilar(reader, out_path=str(cache_path))
    assert r.ok
    assert "cacheado" in r.warnings[0].lower()
    assert load_perfil(str(cache_path)).n(("adultos", "riesgo_sistemico_asa")) == 99


def test_run_perfilar_fallo_sin_cache_falla(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    reader = FakeSheetReader([], fail=True)
    r = orchestrator.run_perfilar(reader, out_path=str(tmp_path / "knowledge" / "perfil_epe.yaml"))
    assert not r.ok


def test_run_perfilar_fallo_cache_con_warnings_vacio_no_crashea(tmp_path, monkeypatch):
    from core.result import AgentResult
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    cache_path = tmp_path / "knowledge" / "perfil_epe.yaml"
    guardar_perfil(Perfil(n_por_celda={("adultos", "riesgo_sistemico_asa"): 99},
                          distribuciones={}, generado_en="2026-07-01"), str(cache_path))
    monkeypatch.setattr(
        orchestrator, "perfilar",
        lambda reader, plantilla: AgentResult.failure([]),
    )
    r = orchestrator.run_perfilar(FakeSheetReader([], fail=True), out_path=str(cache_path))
    assert r.ok
    assert "motivo desconocido" in r.warnings[0]


def test_run_perfilar_fallo_sin_cache_warnings_vacio_no_crashea(tmp_path, monkeypatch):
    from core.result import AgentResult
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(
        orchestrator, "perfilar",
        lambda reader, plantilla: AgentResult.failure([]),
    )
    r = orchestrator.run_perfilar(
        FakeSheetReader([], fail=True),
        out_path=str(tmp_path / "knowledge" / "perfil_epe.yaml"),
    )
    assert not r.ok
    assert "motivo desconocido" in r.warnings[0]


def test_run_propose_produce_candidatos(tmp_path, monkeypatch):
    import orchestrator
    repo_root = Path(__file__).resolve().parent.parent
    (tmp_path / "knowledge").mkdir()
    shutil.copy(repo_root / "knowledge" / "plantilla_epe.yaml",
                tmp_path / "knowledge" / "plantilla_epe.yaml")
    monkeypatch.chdir(tmp_path)
    perfil_path = tmp_path / "perfil.yaml"
    guardar_perfil(Perfil(n_por_celda={}, distribuciones={}, generado_en="2026-07-27",
                          n_conjunto={"discapacidad_intelectual": 40}), str(perfil_path))
    llm = FakeLLMClient(responses=[json.dumps({"score": 8.0, "justificacion": "ok"})])
    pubmed = FakePubMedClient({})
    r = orchestrator.run_propose("knowledge/plantilla_epe.yaml", str(perfil_path), pubmed, llm)
    assert r.ok
    assert len(r.data) >= 1
    assert all(row["candidato"].n_disponible >= 30 for row in r.data)


def test_run_propose_avisa_perfil_precache_multivariado(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    perfil_path = tmp_path / "perfil.yaml"
    guardar_perfil(Perfil(n_por_celda={("adultos", "riesgo_sistemico_asa"): 99},
                          distribuciones={}, generado_en="2026-07-01",
                          n_conjunto={}), str(perfil_path))
    llm = FakeLLMClient(default='{"score": 0, "justificacion": ""}')
    pubmed = FakePubMedClient({})
    r = orchestrator.run_propose("knowledge/plantilla_epe.yaml", str(perfil_path), pubmed, llm)
    assert any("n_conjunto" in w and "perfilar" in w for w in r.warnings)


def test_run_propose_no_avisa_si_perfil_legitimamente_vacio(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    perfil_path = tmp_path / "perfil.yaml"
    guardar_perfil(Perfil(n_por_celda={}, distribuciones={}, generado_en="2026-07-01",
                          n_conjunto={}), str(perfil_path))
    llm = FakeLLMClient(default='{"score": 0, "justificacion": ""}')
    pubmed = FakePubMedClient({})
    r = orchestrator.run_propose("knowledge/plantilla_epe.yaml", str(perfil_path), pubmed, llm)
    assert not any("n_conjunto" in w for w in r.warnings)


def test_orchestrator_main_uso_sin_argumentos(capsys):
    import orchestrator
    assert orchestrator.main([]) == 2
    err = capsys.readouterr().err
    assert "perfilar" in err
    assert "propose" in err


def test_cmd_propose_con_fallo_no_escribe_y_retorna_1(tmp_path, monkeypatch, capsys):
    from core.result import AgentResult
    import orchestrator
    monkeypatch.chdir(tmp_path)
    (tmp_path / "knowledge").mkdir()
    (tmp_path / "knowledge" / "perfil_epe.yaml").write_text("generado_en: '2026-07-24'\n",
                                                             encoding="utf-8")
    monkeypatch.setattr(
        orchestrator, "run_propose",
        lambda *a, **k: AgentResult.failure(["motivo de prueba"]),
    )
    codigo = orchestrator.main(["propose"])
    assert codigo == 1
    salida = capsys.readouterr()
    assert "motivo de prueba" in salida.err
    assert not (tmp_path / "outputs").exists()


def test_cmd_propose_sin_perfil_no_crashea_y_retorna_1(tmp_path, monkeypatch, capsys):
    import orchestrator
    monkeypatch.chdir(tmp_path)
    codigo = orchestrator.main(["propose"])
    assert codigo == 1
    err = capsys.readouterr().err
    assert "perfilar" in err


def test_run_design_encuentra_candidato_y_genera_protocolo(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    _copiar_limitaciones(tmp_path)
    monkeypatch.chdir(tmp_path)
    candidato_data = {
        "id": "riesgo_sistemico_asa_asa3_alto_riesgo_nivel_tratamiento_requerido_adj_farmacoterapia_polifarmacia",
        "eje": "riesgo_sistemico_asa", "subpoblacion": "asa3_alto_riesgo",
        "outcome": "nivel_tratamiento_requerido", "covariables_ajuste": ["farmacoterapia_polifarmacia"],
        "n_disponible": 1350, "novedad": 1.0, "score_llm": 8.0,
    }
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([candidato_data]), encoding="utf-8")
    r = orchestrator.run_design(candidato_data["id"])
    assert r.ok
    assert r.data.candidato_id == candidato_data["id"]


def test_run_design_candidato_no_encontrado(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    _copiar_limitaciones(tmp_path)
    monkeypatch.chdir(tmp_path)
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([]), encoding="utf-8")
    r = orchestrator.run_design("no_existe")
    assert not r.ok


def test_run_design_sin_candidatos_json_falla(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    r = orchestrator.run_design("cualquiera")
    assert not r.ok


def test_cmd_design_escribe_md_y_docx(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    _copiar_limitaciones(tmp_path)
    monkeypatch.chdir(tmp_path)
    candidato_data = {
        "id": "abc", "eje": "riesgo_sistemico_asa", "subpoblacion": "asa3_alto_riesgo",
        "outcome": "nivel_tratamiento_requerido", "covariables_ajuste": ["farmacoterapia_polifarmacia"],
        "n_disponible": 1350, "novedad": 1.0, "score_llm": 8.0,
    }
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([candidato_data]), encoding="utf-8")
    exit_code = orchestrator.main(["design", "abc"])
    assert exit_code == 0
    archivos_md = list((tmp_path / "outputs").glob("*/protocolo.md"))
    archivos_docx = list((tmp_path / "outputs").glob("*/protocolo.docx"))
    assert len(archivos_md) == 1
    assert len(archivos_docx) == 1


def test_run_analyze_encuentra_candidato_y_genera_do(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    candidato_data = {
        "id": "riesgo_sistemico_asa_asa3_alto_riesgo_nivel_tratamiento_requerido_adj_farmacoterapia_polifarmacia",
        "eje": "riesgo_sistemico_asa", "subpoblacion": "asa3_alto_riesgo",
        "outcome": "nivel_tratamiento_requerido", "covariables_ajuste": ["farmacoterapia_polifarmacia"],
        "n_disponible": 1350, "novedad": 1.0, "score_llm": 8.0,
    }
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([candidato_data]), encoding="utf-8")
    r = orchestrator.run_analyze(candidato_data["id"])
    assert r.ok
    assert "ologit nivel_tratamiento_requerido" in r.data


def test_run_analyze_candidato_no_encontrado(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([]), encoding="utf-8")
    r = orchestrator.run_analyze("no_existe")
    assert not r.ok


def test_run_analyze_sin_candidatos_json_falla(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    r = orchestrator.run_analyze("cualquiera")
    assert not r.ok


def test_cmd_analyze_escribe_do(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    monkeypatch.chdir(tmp_path)
    candidato_data = {
        "id": "abc", "eje": "riesgo_sistemico_asa", "subpoblacion": "asa3_alto_riesgo",
        "outcome": "nivel_tratamiento_requerido", "covariables_ajuste": ["farmacoterapia_polifarmacia"],
        "n_disponible": 1350, "novedad": 1.0, "score_llm": 8.0,
    }
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([candidato_data]), encoding="utf-8")
    exit_code = orchestrator.main(["analyze", "abc"])
    assert exit_code == 0
    archivos = list((tmp_path / "outputs").glob("*/analisis.do"))
    assert len(archivos) == 1


def test_run_report_encuentra_candidato_y_genera_articulo(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    _copiar_limitaciones(tmp_path)
    monkeypatch.chdir(tmp_path)
    candidato_data = {
        "id": "riesgo_sistemico_asa_asa3_alto_riesgo_nivel_tratamiento_requerido_adj_farmacoterapia_polifarmacia",
        "eje": "riesgo_sistemico_asa", "subpoblacion": "asa3_alto_riesgo",
        "outcome": "nivel_tratamiento_requerido", "covariables_ajuste": ["farmacoterapia_polifarmacia"],
        "n_disponible": 1350, "novedad": 1.0, "score_llm": 8.0,
    }
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([candidato_data]), encoding="utf-8")
    xlsx_path = _resultados_xlsx_valido(tmp_path)
    r = orchestrator.run_report(candidato_data["id"], xlsx_path)
    assert r.ok
    assert "riesgo_sistemico_asa" in r.data.resultados


def test_run_report_candidato_no_encontrado(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    _copiar_limitaciones(tmp_path)
    monkeypatch.chdir(tmp_path)
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([]), encoding="utf-8")
    xlsx_path = _resultados_xlsx_valido(tmp_path)
    r = orchestrator.run_report("no_existe", xlsx_path)
    assert not r.ok


def test_run_report_xlsx_invalido_falla(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    _copiar_limitaciones(tmp_path)
    monkeypatch.chdir(tmp_path)
    candidato_data = {
        "id": "abc", "eje": "riesgo_sistemico_asa", "subpoblacion": "asa3_alto_riesgo",
        "outcome": "nivel_tratamiento_requerido", "covariables_ajuste": [],
        "n_disponible": 100, "novedad": 1.0, "score_llm": 8.0,
    }
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([candidato_data]), encoding="utf-8")
    wb = openpyxl.Workbook()
    wb.active.title = "descriptivos"
    xlsx_path = tmp_path / "resultados_incompleto.xlsx"
    wb.save(xlsx_path)
    r = orchestrator.run_report("abc", str(xlsx_path))
    assert not r.ok


def test_cmd_report_escribe_articulo(tmp_path, monkeypatch):
    import orchestrator
    _copiar_plantilla(tmp_path)
    _copiar_limitaciones(tmp_path)
    monkeypatch.chdir(tmp_path)
    candidato_data = {
        "id": "abc", "eje": "riesgo_sistemico_asa", "subpoblacion": "asa3_alto_riesgo",
        "outcome": "nivel_tratamiento_requerido", "covariables_ajuste": ["farmacoterapia_polifarmacia"],
        "n_disponible": 1350, "novedad": 1.0, "score_llm": 8.0,
    }
    out_dir = tmp_path / "outputs" / "20260728-000000"
    out_dir.mkdir(parents=True)
    (out_dir / "candidatos.json").write_text(json.dumps([candidato_data]), encoding="utf-8")
    xlsx_path = _resultados_xlsx_valido(tmp_path)
    exit_code = orchestrator.main(["report", "abc", xlsx_path])
    assert exit_code == 0
    archivos = list((tmp_path / "outputs").glob("*/articulo.md"))
    assert len(archivos) == 1
