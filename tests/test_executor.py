import openpyxl

from agents.executor import parsear_resultados


def _libro_valido(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "descriptivos"
    ws.append([None, "nivel_tratamiento_requerido"])
    ws.append(["b", 2.34])
    ws.append(["ll", 2.10])
    ws.append(["ul", 2.58])

    ws2 = wb.create_sheet("modelo")
    # Fila de ecuacion repetida (nombre del outcome en ambas columnas) + fila 2
    # con los terminos reales -- simula la salida real de un comando eclass
    # (ologit/mlogit/logistic/regress) via `putexcel ... matrix(r(table))`.
    ws2.append([None, "nivel_tratamiento_requerido", "nivel_tratamiento_requerido"])
    ws2.append([None, "riesgo_sistemico_asa", "_cons"])
    ws2.append(["b", 1.87, 0.5])
    ws2.append(["ll", 1.20, 0.3])
    ws2.append(["ul", 2.91, 0.8])
    ws2.append(["pvalue", 0.003, 0.01])

    ws3 = wb.create_sheet("bivariado_farmacoterapia_polifarmacia")
    ws3.append([None, "1", "2"])
    ws3.append(["b", 3.1, 1.5])
    ws3.append(["ll", 2.0, 0.9])
    ws3.append(["ul", 4.2, 2.1])

    path = tmp_path / "resultados.xlsx"
    wb.save(path)
    return path


def test_parsear_resultados_feliz(tmp_path):
    path = _libro_valido(tmp_path)
    r = parsear_resultados(str(path))
    assert r.ok
    assert r.data["descriptivos"][0]["efecto"] == 2.34
    terminos_modelo = {t["termino"]: t for t in r.data["modelo"]}
    assert terminos_modelo["riesgo_sistemico_asa"]["efecto"] == 1.87
    assert terminos_modelo["riesgo_sistemico_asa"]["p"] == 0.003
    assert r.data["bivariado"]["farmacoterapia_polifarmacia"][0]["efecto"] == 3.1


def test_parsear_resultados_hoja_faltante(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "descriptivos"
    ws.append([None, "x"])
    ws.append(["b", 1.0])
    ws.append(["ll", 0.5])
    ws.append(["ul", 1.5])
    path = tmp_path / "resultados.xlsx"
    wb.save(path)
    r = parsear_resultados(str(path))
    assert not r.ok
    assert "modelo" in r.warnings[0]


def test_parsear_resultados_fila_requerida_faltante(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "descriptivos"
    ws.append([None, "x"])
    ws.append(["b", 1.0])
    ws.append(["ll", 0.5])
    # falta "ul"
    ws2 = wb.create_sheet("modelo")
    ws2.append([None, "x"])
    ws2.append(["b", 1.0])
    ws2.append(["ll", 0.5])
    ws2.append(["ul", 1.5])
    path = tmp_path / "resultados.xlsx"
    wb.save(path)
    r = parsear_resultados(str(path))
    assert not r.ok
    assert "descriptivos" in r.warnings[0]
    assert "'ul'" in r.warnings[0]


def test_parsear_resultados_celda_no_numerica(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "descriptivos"
    ws.append([None, "x"])
    ws.append(["b", "no_numero"])
    ws.append(["ll", 0.5])
    ws.append(["ul", 1.5])
    ws2 = wb.create_sheet("modelo")
    ws2.append([None, "x"])
    ws2.append(["b", 1.0])
    ws2.append(["ll", 0.5])
    ws2.append(["ul", 1.5])
    path = tmp_path / "resultados.xlsx"
    wb.save(path)
    r = parsear_resultados(str(path))
    assert not r.ok
    assert "celda no numérica" in r.warnings[0]


def test_parsear_resultados_bivariado_mal_formado_se_omite(tmp_path):
    path = _libro_valido(tmp_path)
    wb = openpyxl.load_workbook(path)
    ws_malo = wb.create_sheet("bivariado_otro")
    ws_malo.append([None, "1"])
    ws_malo.append(["b", 1.0])
    # falta ll/ul
    wb.save(path)
    r = parsear_resultados(str(path))
    assert r.ok
    assert "otro" not in r.data["bivariado"]
    assert any("bivariado_otro" in w for w in r.warnings)
