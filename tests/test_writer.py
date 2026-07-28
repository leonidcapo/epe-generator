import openpyxl

from agents.bias_auditor import load_limitaciones
from agents.executor import parsear_resultados
from agents.novelty_checker import Candidato
from agents.writer import redactar_articulo, redactar_resultados
from core.knowledge import load_plantilla
from core.llm_client import FakeLLMClient


def _plantilla():
    return load_plantilla("knowledge/plantilla_epe.yaml")


def _candidato():
    return Candidato(eje="riesgo_sistemico_asa", subpoblacion="asa3_alto_riesgo",
                     outcome="nivel_tratamiento_requerido",
                     covariables_ajuste=("farmacoterapia_polifarmacia",), n_disponible=1350)


def _tablas():
    return {
        "descriptivos": [{"termino": "nivel_tratamiento_requerido", "efecto": 2.34,
                          "ic_inf": 2.10, "ic_sup": 2.58, "p": None}],
        "modelo": [
            {"termino": "riesgo_sistemico_asa", "efecto": 1.87, "ic_inf": 1.20,
             "ic_sup": 2.91, "p": 0.003},
            {"termino": "_cons", "efecto": 0.5, "ic_inf": 0.3, "ic_sup": 0.8, "p": 0.01},
        ],
        "bivariado": {},
    }


def test_redactar_resultados_excluye_cons():
    c = _candidato()
    texto = redactar_resultados(_tablas(), c)
    assert "riesgo_sistemico_asa" in texto
    assert "1,87" in texto
    assert "_cons" not in texto


def test_redactar_resultados_traduce_hoja_bivariado_a_covariable_real():
    from agents.statistician import mapeo_hojas_bivariado
    c = Candidato(eje="riesgo_sistemico_asa", subpoblacion="asa3_alto_riesgo",
                 outcome="nivel_tratamiento_requerido",
                 covariables_ajuste=("farmacoterapia_polifarmacia",), n_disponible=1350)
    hoja_farmaco_completa = mapeo_hojas_bivariado([c.eje, *c.covariables_ajuste])["farmacoterapia_polifarmacia"]
    hoja_farmaco_stripped = hoja_farmaco_completa[len("bivariado_"):]  # lo que executor.py realmente produce
    tablas = {
        "descriptivos": [], "modelo": [],
        "bivariado": {hoja_farmaco_stripped: [{"termino": "1", "efecto": 3.1, "ic_inf": 2.0, "ic_sup": 4.2, "p": None}]},
    }
    texto = redactar_resultados(tablas, c)
    assert "farmacoterapia_polifarmacia = 1:" in texto
    # el fragmento truncado no debe aparecer como token propio (solo como substring
    # del nombre real ya traducido, que sí lo contiene por coincidencia de prefijo)
    assert f"{hoja_farmaco_stripped} =" not in texto


def test_redactar_resultados_traduce_correctamente_via_executor_real(tmp_path):
    """Prueba de integracion real: escribe un .xlsx con el MISMO nombre de hoja
    truncado que agents/statistician.py generaria, lo parsea con el executor real
    (no un dict armado a mano), y confirma que redactar_resultados muestra el
    nombre real de la covariable, no el fragmento truncado."""
    from agents.statistician import mapeo_hojas_bivariado

    c = Candidato(eje="riesgo_sistemico_asa", subpoblacion="asa3_alto_riesgo",
                 outcome="nivel_tratamiento_requerido",
                 covariables_ajuste=("farmacoterapia_polifarmacia",), n_disponible=1350)
    hoja_farmaco = mapeo_hojas_bivariado([c.eje, *c.covariables_ajuste])["farmacoterapia_polifarmacia"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "descriptivos"
    ws.append([None, "nivel_tratamiento_requerido"])
    ws.append(["b", 2.34])
    ws.append(["ll", 2.10])
    ws.append(["ul", 2.58])
    ws2 = wb.create_sheet("modelo")
    ws2.append([None, "riesgo_sistemico_asa"])
    ws2.append(["b", 1.87])
    ws2.append(["ll", 1.20])
    ws2.append(["ul", 2.91])
    ws3 = wb.create_sheet(hoja_farmaco)  # nombre de hoja EXACTAMENTE como lo escribiria statistician.py
    ws3.append([None, "1"])
    ws3.append(["b", 3.1])
    ws3.append(["ll", 2.0])
    ws3.append(["ul", 4.2])
    path = tmp_path / "resultados.xlsx"
    wb.save(path)

    resultado = parsear_resultados(str(path))
    assert resultado.ok
    texto = redactar_resultados(resultado.data, c)
    assert "farmacoterapia_polifarmacia = 1:" in texto


def test_redactar_articulo_degrada_sin_llm():
    r = redactar_articulo(_candidato(), _plantilla(), _tablas(), [], None)
    assert r.ok
    assert r.data.prosa_post["discusion"] == "[pendiente: LLM no disponible]"
    assert r.data.prosa_ante["introduccion"] == "[prosa pendiente: LLM no disponible]"


def test_redactar_articulo_con_llm_disponible():
    llm = FakeLLMClient(responses=["Texto redactado en pasado."])
    r = redactar_articulo(_candidato(), _plantilla(), _tablas(), [], llm)
    assert r.ok
    assert r.data.prosa_post["discusion"] == "Texto redactado en pasado."


def test_redactar_articulo_marca_cifra_inventada():
    llm = FakeLLMClient(responses=["El efecto fue de 9.99 con IC muy amplio."])
    r = redactar_articulo(_candidato(), _plantilla(), _tablas(), [], llm)
    assert r.data.prosa_post["discusion"] == "[sección pendiente: cifra no verificable]"
    assert any("cifras no verificables" in w for w in r.warnings)


def test_redactar_articulo_detecta_lenguaje_causal():
    lims = load_limitaciones("knowledge/limitaciones_epe.yaml")
    llm = FakeLLMClient(responses=["El riesgo sistémico causó el nivel de tratamiento."])
    r = redactar_articulo(_candidato(), _plantilla(), _tablas(), lims, llm)
    assert any("Lenguaje causal" in w for w in r.warnings)


def test_redactar_articulo_candidato_id_coincide_con_protocolo():
    from agents.novelty_checker import candidato_id
    c = _candidato()
    llm = FakeLLMClient(responses=["Texto en pasado."])
    r = redactar_articulo(c, _plantilla(), _tablas(), [], llm)
    assert r.data.candidato_id == candidato_id(c)
